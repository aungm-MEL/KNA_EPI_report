"""
build_kna_clean.py
------------------
Reads  : KNA Child vaccination.xlsx  (same folder as this script)
Writes : KNA_clean.xlsx              (same folder)

Changes applied
---------------
  1. Keeps    ChNa  column.
  2. Adds     children_code column (= Reg_ + "_" + Site).
  3. Adds     Full Dose quarter columns for Q1 2025 – Q4 2026.
  4. Adds     completion_status_2023 column (annual formula, Jan–Dec 2023).

completion_status_2023 differences from quarterly formula
---------------------------------------------------------
  Age column         : M1A  (col AI = age when Measles1 was given)
                       NOT FSD A. The 2023 formula was written for a file where
                       column AI maps to M1A in the source data.
  Dose-in-year check : 2023-01-01 – 2023-12-21
  MAX boundary       : 2023-12-31  (different from dose-check end)
  1-5 age group      : M1A 12-59 months OR M1A == 111 (special age code)
----------------------------------------------------------
  "U1 Completed in Qx YYYY"  -> FSD A 0-11 months (0 is valid month age)
    AND BCG+Pe1-3+OPV1-3+MM1 all given (status 1/2/5)
    AND at least one dose date falls within the quarter AND status 2/5
    AND MAX(Pe3D, OP3D, MM1D) <= quarter end date

  "1-5 Completed in Qx YYYY" -> FSD A 12-59 months
    AND Pe1-3+OPV1-3+MM1 all given (status 1/2/5)
    AND at least one dose date falls within the quarter AND status 2/5
    AND MAX(Pe3D, OP3D, MM1D) <= quarter end date

  ""  -> all other cases

Quarter date boundaries
-----------------------
  Q1 2025 : 2024-12-21 – 2025-03-20
  Q2 2025 : 2025-03-21 – 2025-06-20
  Q3 2025 : 2025-06-21 – 2025-09-20
  Q4 2025 : 2025-09-21 – 2025-12-20
  Q1 2026 : 2025-12-21 – 2026-03-20
  Q2 2026 : 2026-03-21 – 2026-06-20
  Q3 2026 : 2026-06-21 – 2026-09-20
  Q4 2026 : 2026-09-21 – 2026-12-20
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill

VALID_STATUS  = {1, 2, 5}  # vaccine was given (any period)
PERIOD_STATUS = {2, 5}     # vaccine was given specifically in the measured period

# Quarter definitions: (output column name, start date, end date)
QUARTERS = [
  ("Full Dose in Q1 2024", pd.Timestamp("2023-12-21"), pd.Timestamp("2024-03-20")),
  ("Full Dose in Q2 2024", pd.Timestamp("2024-03-21"), pd.Timestamp("2024-06-20")),
  ("Full Dose in Q3 2024", pd.Timestamp("2024-06-21"), pd.Timestamp("2024-09-20")),
  ("Full Dose in Q4 2024", pd.Timestamp("2024-09-21"), pd.Timestamp("2024-12-20")),
  ("Full Dose in Q1 2025", pd.Timestamp("2024-12-21"), pd.Timestamp("2025-03-20")),
  ("Full Dose in Q2 2025", pd.Timestamp("2025-03-21"), pd.Timestamp("2025-06-20")),
  ("Full Dose in Q3 2025", pd.Timestamp("2025-06-21"), pd.Timestamp("2025-09-20")),
  ("Full Dose in Q4 2025", pd.Timestamp("2025-09-21"), pd.Timestamp("2025-12-20")),
  ("Full Dose in Q1 2026", pd.Timestamp("2025-12-21"), pd.Timestamp("2026-03-20")),
  ("Full Dose in Q2 2026", pd.Timestamp("2026-03-21"), pd.Timestamp("2026-06-20")),
  ("Full Dose in Q3 2026", pd.Timestamp("2026-06-21"), pd.Timestamp("2026-09-20")),
  ("Full Dose in Q4 2026", pd.Timestamp("2026-09-21"), pd.Timestamp("2026-12-20")),
]


def _status_ok(val) -> bool:
  """True if the status value indicates the vaccine was given."""
  try:
    if isinstance(val, str):
      val = val.strip()
    return int(float(val)) in VALID_STATUS
  except (TypeError, ValueError):
    return False


def _in_period(date_val, status_val, start: pd.Timestamp, end: pd.Timestamp) -> bool:
  """True if the dose date falls within [start, end] AND status is 2 or 5."""
  try:
    if pd.isna(date_val) or pd.isna(status_val):
      return False
    if isinstance(status_val, str):
      status_val = status_val.strip()
    status_num = int(float(status_val))

    # Handle Excel serial dates if present as numbers.
    if isinstance(date_val, (int, float)):
      dt = pd.Timestamp("1899-12-30") + pd.to_timedelta(float(date_val), unit="D")
    else:
      dt = pd.Timestamp(date_val)

    return start <= dt <= end and status_num in PERIOD_STATUS
  except (TypeError, ValueError):
    return False


def _max_date(*date_vals) -> pd.Timestamp:
  """Return the maximum of the given date values, ignoring NaT/NaN."""
  dates = []
  for d in date_vals:
    try:
      if pd.notna(d):
        dates.append(pd.Timestamp(d))
    except (TypeError, ValueError):
      pass
  return max(dates) if dates else pd.NaT


def build_indicators_2025() -> "pd.DataFrame":
    """
    Build the 2025_indicators summary sheet.

    Structure: 30 rows (3 years × 10 indicators), 24 columns.
    Targets are taken from the pre-defined programme plan.
    Achievement counts (U1 Male/Female, 1-5 Male/Female) are left blank (NaN)
    and will be filled in a later stage.
    2024 rows are intentionally empty (NaN targets) as per reporting convention.
    """
    INDICATORS = [
        "Penta3 under 1-yr-old",
        "MMR1 under 1-yr-old",
        "Penta1 under 5-yr-old",
        "Penta3 under 5-yr-old",
        "MMR1 under 5-yr-old",
        "MMR2 under 5-yr-old",
        "Full dose under 5-yr-old",
        "At least one dose under 5-yr-old",
        "Td ALOD",
        "Td Two Doses",
    ]

    # Quarterly targets [Q1, Q2, Q3, Q4] — None means no target set for that quarter
    TARGETS = {
        2024: {ind: [None, None, None, None] for ind in INDICATORS},
        2025: {
            "Penta3 under 1-yr-old":           [1600/3,    1600/3,    1600/3,    1600/3],
            "MMR1 under 1-yr-old":             [1400/3,    1400/3,    1400/3,    1400/3],
            "Penta1 under 5-yr-old":           [4250/3,    4250/3,    4250/3,    4250/3],
            "Penta3 under 5-yr-old":           [3400/3,    3400/3,    3400/3,    3400/3],
            "MMR1 under 5-yr-old":             [1275.0,    1275.0,    1275.0,    1275.0],
            "MMR2 under 5-yr-old":             [1020.0,    1020.0,    1020.0,    1020.0],
            "Full dose under 5-yr-old":        [6025/6,    6025/6,    6025/6,    6025/6],
            "At least one dose under 5-yr-old":[7000/3,    7000/3,    7000/3,    7000/3],
            "Td ALOD":                         [2000/3,    2000/3,    2000/3,    2000/3],
            "Td Two Doses":                    [2725/6,    2725/6,    2725/6,    2725/6],
        },
        2026: {
            "Penta3 under 1-yr-old":           [4078/3,    4078/3,    None,       None],
            "MMR1 under 1-yr-old":             [2848/3,    2848/3,    None,       None],
            "Penta1 under 5-yr-old":           [3170.0,    3170.0,    None,       None],
            "Penta3 under 5-yr-old":           [8560/3,    8560/3,    None,       None],
            "MMR1 under 5-yr-old":             [8662/3,    8662/3,    None,       None],
            "MMR2 under 5-yr-old":             [7796/3,    7796/3,    None,       None],
            "Full dose under 5-yr-old":        [9838/3,    9838/3,    None,       None],
            "At least one dose under 5-yr-old":[17572/3,   17572/3,   None,       None],
            "Td ALOD":                         [4144/3,    4144/3,    None,       None],
            "Td Two Doses":                    [2900/3,    2900/3,    None,       None],
        },
    }

    rows = []
    for year in (2024, 2025, 2026):
        year_targets = TARGETS[year]
        for ind in INDICATORS:
            q = year_targets[ind]
            rows.append({
                "Period":        year,
                "Organization":  "KNA",
                "Project Name":  "REACH-KK",
                "indicator":     ind,
                "Q1 Target":     q[0],
                "Q1 U1 Male":    None,
                "Q1 U1 Female":  None,
                "Q1 1-5 Male":   None,
                "Q1 1-5 Female": None,
                "Q2 Target":     q[1],
                "Q2 U1 Male":    None,
                "Q2 U1 Female":  None,
                "Q2 1-5 Male":   None,
                "Q2 1-5 Female": None,
                "Q3 Target":     q[2],
                "Q3 U1 Male":    None,
                "Q3 U1 Female":  None,
                "Q3 1-5 Male":   None,
                "Q3 1-5 Female": None,
                "Q4 Target":     q[3],
                "Q4 U1 Male":    None,
                "Q4 U1 Female":  None,
                "Q4 1-5 Male":   None,
                "Q4 1-5 Female": None,
            })

    return pd.DataFrame(rows)


def full_dose_quarter(
    row,
    start: pd.Timestamp,
    end: pd.Timestamp,
    label: str,
    max_end: pd.Timestamp = None,
    extra_ages_15: frozenset = frozenset(),
    age_col: str = "FSD A",
) -> str:
  """
  Evaluate Full Dose completion for a given period.

  Parameters
  ----------
  start / end     : dose-in-period date window (inclusive).
  label           : text embedded in the result string.
  max_end         : upper bound for MAX(Pe3D, OP3D, MM1D) check.
                    Defaults to `end` when None.
  extra_ages_15   : additional age values (besides 12-59) that qualify
                    for the 1-5 age group (e.g. frozenset({111})).
  age_col         : column to use for the age check (default: FSD A).
                    Use 'M1A' for formulas written against column AI.
  """
  age = row[age_col]
  if pd.isna(age):
    return ""
  try:
    age = float(age)
  except (TypeError, ValueError):
    return ""

  # Match Excel: MAX(blank,blank,blank)=0 which is always <= any target date.
  _max_end = end if max_end is None else max_end
  max_d    = _max_date(row["Pe3D"], row["OP3D"], row["MM1D"])
  max_ok   = pd.isna(max_d) or max_d <= _max_end

  # Condition 1: U1 (FSD A 0-11 months; 0 is valid month age)
  if 0 <= age <= 11:
    status_u1 = (
      _status_ok(row["BC_C"]) and _status_ok(row["PE1C"]) and
      _status_ok(row["PE2C"]) and _status_ok(row["PE3C"]) and
      _status_ok(row["OP1C"]) and _status_ok(row["OP2C"]) and
      _status_ok(row["OP3C"]) and _status_ok(row["MM1C"])
    )
    any_period_u1 = (
      _in_period(row["BCGD"], row["BC_C"], start, end) or
      _in_period(row["Pe1D"], row["PE1C"], start, end) or
      _in_period(row["Pe2D"], row["PE2C"], start, end) or
      _in_period(row["Pe3D"], row["PE3C"], start, end) or
      _in_period(row["OP1D"], row["OP1C"], start, end) or
      _in_period(row["OP2D"], row["OP2C"], start, end) or
      _in_period(row["OP3D"], row["OP3C"], start, end) or
      _in_period(row["MM1D"], row["MM1C"], start, end) or
      _in_period(row["MM2D"], row["MM2C"], start, end)
    )
    if status_u1 and any_period_u1 and max_ok:
      return f"U1 Completed in {label}"

  # Condition 2: 1-5 years (FSD A 12-59 months, or any extra_ages_15 code)
  elif (11 < age <= 59) or (age in extra_ages_15):
    status_15 = (
      _status_ok(row["PE1C"]) and _status_ok(row["PE2C"]) and
      _status_ok(row["PE3C"]) and _status_ok(row["OP1C"]) and
      _status_ok(row["OP2C"]) and _status_ok(row["OP3C"]) and
      _status_ok(row["MM1C"])
    )
    any_period_15 = (
      _in_period(row["Pe1D"], row["PE1C"], start, end) or
      _in_period(row["Pe2D"], row["PE2C"], start, end) or
      _in_period(row["Pe3D"], row["PE3C"], start, end) or
      _in_period(row["OP1D"], row["OP1C"], start, end) or
      _in_period(row["OP2D"], row["OP2C"], start, end) or
      _in_period(row["OP3D"], row["OP3C"], start, end) or
      _in_period(row["MM1D"], row["MM1C"], start, end) or
      _in_period(row["MM2D"], row["MM2C"], start, end)
    )
    if status_15 and any_period_15 and max_ok:
      return f"1-5 Completed in {label}"

  return ""

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent


def _path_from_env(env_name: str, default_path: Path) -> Path:
  raw = os.getenv(env_name, "").strip()
  if not raw:
    return default_path
  return Path(raw).expanduser().resolve()


SRC    = _path_from_env("KNA_CHILD_SRC", BASE_DIR / "KNA Child vaccination.xlsx")
TD_SRC = _path_from_env("KNA_TD_SRC", BASE_DIR / "KNA Td Vaccination.xlsx")
DST    = _path_from_env("KNA_CLEAN_DST", BASE_DIR / "KNA_clean.xlsx")

# Fallback to current working directory when no env override is provided.
if not SRC.exists() and SRC.name == "KNA Child vaccination.xlsx":
  cwd_src = Path.cwd() / SRC.name
  if cwd_src.exists():
    SRC = cwd_src

if not TD_SRC.exists() and TD_SRC.name == "KNA Td Vaccination.xlsx":
  cwd_td = Path.cwd() / TD_SRC.name
  if cwd_td.exists():
    TD_SRC = cwd_td

if not SRC.exists():
  raise FileNotFoundError(f"Child source file not found: {SRC}")
if not TD_SRC.exists():
  raise FileNotFoundError(f"Td source file not found: {TD_SRC}")

DST.parent.mkdir(parents=True, exist_ok=True)

# ── Read ───────────────────────────────────────────────────────────────────────
print(f"Reading  : {SRC}")
df = pd.read_excel(SRC, engine="openpyxl")
print(f"  Rows   : {len(df)}")
print(f"  Columns: {len(df.columns)}")

print(f"Reading  : {TD_SRC} [AN Td]")
td_df = pd.read_excel(TD_SRC, sheet_name="AN Td", engine="openpyxl")
print(f"  Td rows   : {len(td_df)}")
print(f"  Td columns: {len(td_df.columns)}")

def _combine_pw_code_td(row, reg_col: str, site_col: str):
  reg = str(row[reg_col]).strip() if pd.notna(row[reg_col]) else ""
  site = str(row[site_col]).strip() if pd.notna(row[site_col]) else ""
  if reg and site:
    return f"{reg}_{site}"
  return reg or site or None


td_reg_col = None
for candidate in ["Reg_", "Regn", "Reg", "Reg No", "RegNo"]:
  if candidate in td_df.columns:
    td_reg_col = candidate
    break

td_site_col = "Site" if "Site" in td_df.columns else None

if td_reg_col and td_site_col:
  print(f"Building 'pw_code' from {td_reg_col} + {td_site_col}...")
  td_df["pw_code"] = td_df.apply(lambda r: _combine_pw_code_td(r, td_reg_col, td_site_col), axis=1)

  # Insert pw_code right after site column for readability.
  site_pos_td = td_df.columns.get_loc(td_site_col)
  td_cols = td_df.columns.tolist()
  td_cols.remove("pw_code")
  td_cols.insert(site_pos_td + 1, "pw_code")
  td_df = td_df[td_cols]

td_code_col = "pw_code" if "pw_code" in td_df.columns else None
if td_code_col is None:
  for candidate in ["PW_code"]:
    if candidate in td_df.columns:
      td_code_col = candidate
      break

if td_code_col is None:
  print("pw_code duplicate highlight skipped: pw_code/PW_code column not found in Td sheet.")
  td_dup_mask = pd.Series(False, index=td_df.index)
else:
  td_dup_mask = (
    td_df[td_code_col].notna()
    & (td_df[td_code_col].astype(str).str.strip() != "")
    & td_df[td_code_col].duplicated(keep=False)
  )
  print(f"{td_code_col} duplicates: {int(td_dup_mask.sum())}")

# ── Validate required columns ──────────────────────────────────────────────────
required = [
  "Reg_", "Site", "FSD A",
  "BCGD", "BC_C",
  "Pe1D", "PE1C", "Pe2D", "PE2C", "Pe3D", "PE3C",
  "OP1D", "OP1C", "OP2D", "OP2C", "OP3D", "OP3C",
  "MM1D", "MM1C", "MM2D", "MM2C",
]
missing  = [c for c in required if c not in df.columns]
if missing:
    raise ValueError(f"Required columns not found in source file: {missing}\n"
                     f"Available columns: {df.columns.tolist()}")

# ── Build children_code (Reg_ + Site) ────────────────────────────────────────
def _combine_children_code(row):
  reg = str(row["Reg_"]).strip() if pd.notna(row["Reg_"]) else ""
  site = str(row["Site"]).strip() if pd.notna(row["Site"]) else ""
  if reg and site:
    return f"{reg}_{site}"
  return reg or site or None


df["children_code"] = df.apply(_combine_children_code, axis=1)

# Insert children_code right after Site for readability.
site_pos = df.columns.get_loc("Site")
cols = df.columns.tolist()
cols.remove("children_code")
cols.insert(site_pos + 1, "children_code")
df = df[cols]

dup_mask = (
  df["children_code"].notna()
  & (df["children_code"].astype(str).str.strip() != "")
  & df["children_code"].duplicated(keep=False)
)
print(f"children_code duplicates: {int(dup_mask.sum())}")

# -- Add Full Dose columns for 2025 and 2026 quarters only ---------------------
for col_name, q_start, q_end in QUARTERS:
    # Only add quarterly columns for 2025 and 2026
    if q_start >= pd.Timestamp("2024-12-21"):
        print(f"\nCalculating '{col_name}'...")
        q_label = col_name.replace("Full Dose in ", "")
        df[col_name] = df.apply(
            lambda r, s=q_start, e=q_end, l=q_label: full_dose_quarter(r, s, e, l),
            axis=1,
        )
        counts = df[col_name].value_counts(dropna=False)
        print(counts.to_string())

# -- Add completion_status_2023 ------------------------------------------------
# Dose-in-year window : 2023-01-01 – 2023-12-21
# MAX boundary        : 2023-12-31  (different from dose-check end)
# Extra 1-5 age code  : FSD A == 111
print("\nCalculating 'completion_status_2023'...")
df["completion_status_2023"] = df.apply(
    lambda r: full_dose_quarter(
        r,
        start=pd.Timestamp("2023-01-01"),
        end=pd.Timestamp("2023-12-21"),
        label="2023",
        max_end=pd.Timestamp("2023-12-31"),
        extra_ages_15=frozenset({111}),
        age_col="M1A",  # Excel col AI = M1A in source data
    ),
    axis=1,
)
counts = df["completion_status_2023"].value_counts(dropna=False)
print(counts.to_string())

# -- Add completion_status_2024 (annual) --------------------------------------
# Dose-in-year window : 2024-01-01 - 2024-12-21
# MAX boundary        : 2024-12-31  (different from dose-check end)
# Extra 1-5 age code  : M1A == 111
print("\nCalculating 'completion_status_2024' (annual)...")
df["completion_status_2024"] = df.apply(
  lambda r: full_dose_quarter(
    r,
    start=pd.Timestamp("2024-01-01"),
    end=pd.Timestamp("2024-12-21"),
    label="2024",
    max_end=pd.Timestamp("2024-12-31"),
    extra_ages_15=frozenset({111}),
    age_col="M1A",  # Excel col AI = M1A in source data
  ),
  axis=1,
)
counts = df["completion_status_2024"].value_counts(dropna=False)
print(counts.to_string())

# -- Add completion_status_2024_Q1 to Q4 (quarterlies, same formula as annual) --
quarterly_2024 = [
  ("completion_status_2024_Q1", pd.Timestamp("2023-12-21"), pd.Timestamp("2024-03-20"), "Q1 2024"),
  ("completion_status_2024_Q2", pd.Timestamp("2024-03-21"), pd.Timestamp("2024-06-20"), "Q2 2024"),
  ("completion_status_2024_Q3", pd.Timestamp("2024-06-21"), pd.Timestamp("2024-09-20"), "Q3 2024"),
  ("completion_status_2024_Q4", pd.Timestamp("2024-09-21"), pd.Timestamp("2024-12-20"), "Q4 2024"),
]
for col, q_start, q_end, q_label in quarterly_2024:
  print(f"\nCalculating '{col}'...")
  df[col] = df.apply(
    lambda r, s=q_start, e=q_end, l=q_label: full_dose_quarter(
      r,
      start=s,
      end=e,
      label=l,
      max_end=pd.Timestamp("2024-12-31"),
      extra_ages_15=frozenset({111}),
      age_col="M1A",
    ),
    axis=1,
  )
  counts = df[col].value_counts(dropna=False)
  print(counts.to_string())

def build_td_alod() -> "pd.DataFrame":
    """
    Td_alod sheet: annual Td ALOD target + achievement (blank).
    3 data rows (2024, 2025, 2026).
    """
    rows = [
        {"Period": 2024, "Organization": "KNA", "Project Name": "REACH-KK",
         "Indicators": "Td ALOD", "Annual Target": None, "Annual Achievement": None},
        {"Period": 2025, "Organization": "KNA", "Project Name": "REACH-KK",
         "Indicators": "Td ALOD", "Annual Target": 2666.0, "Annual Achievement": None},
        {"Period": 2026, "Organization": "KNA", "Project Name": "REACH-KK",
         "Indicators": "Td ALOD", "Annual Target": 5800 / 3, "Annual Achievement": None},
    ]
    return pd.DataFrame(rows)


def build_alod_cummu() -> "pd.DataFrame":
    """
    ALOD_cummu sheet: annual At-least-one-dose target + gender/age breakdown (blank).
    3 data rows (2024, 2025, 2026).
    """
    rows = [
        {"Period": 2024, "Organization": "KNA", "Project Name": "REACH-KK",
         "indicator": "At least one dose under 5-yr-old",
         "Annual Target": 9000.0,
         "Annual U1 Male": None, "Annaul U1 Female": None,
         "Annual 1-5 Male": None, "Annual 1-5 Female": None},
        {"Period": 2025, "Organization": "KNA", "Project Name": "REACH-KK",
         "indicator": "At least one dose under 5-yr-old",
         "Annual Target": 9334.0,
         "Annual U1 Male": None, "Annaul U1 Female": None,
         "Annual 1-5 Male": None, "Annual 1-5 Female": None},
        {"Period": 2026, "Organization": "KNA", "Project Name": "REACH-KK",
         "indicator": "At least one dose under 5-yr-old",
         "Annual Target": 35144 / 3,
         "Annual U1 Male": None, "Annaul U1 Female": None,
         "Annual 1-5 Male": None, "Annual 1-5 Female": None},
    ]
    return pd.DataFrame(rows)


def build_idp() -> "pd.DataFrame":
    """
    IDP sheet: quarterly IDP/non-IDP Male/Female breakdown (all blank achievements).
    3 data rows (2024, 2025, 2026). Indicator: Penta1 under 5-yr-old.
    """
    quarters = ["Q1", "Q2", "Q3", "Q4"]
    cols_order = ["Period", "Organization", "Project Name", "indicator"]
    for q in quarters:
        cols_order += [f"{q} IDP Male", f"{q} IDP Female",
                       f"{q} non-IDP Male", f"{q} non-IDP Female"]

    rows = []
    for year in (2024, 2025, 2026):
        row = {"Period": year, "Organization": "KNA", "Project Name": "REACH_KK",
               "indicator": "Penta1 under 5-yr-old"}
        for q in quarters:
            for col in [f"{q} IDP Male", f"{q} IDP Female",
                        f"{q} non-IDP Male", f"{q} non-IDP Female"]:
                row[col] = None
        rows.append(row)
    return pd.DataFrame(rows, columns=cols_order)


def build_td2_indicator() -> "pd.DataFrame":
    """
    Td2_indicator sheet: quarterly Td Two Doses target + achievement (blank).
    4 data rows (2023, 2024, 2025, 2026).
    2025: Q1-Q4 target = 2725/6 each.
    2026: Q1-Q2 target = 2900/3 each; Q3-Q4 = blank.
    """
    rows = [
        {"Period": 2023, "Organization": "KNA", "Project Name": "REACH-KK",
         "Indicators": "Td Two Doses",
         "Q1 Target": None, "Q1 Achievement": None,
         "Q2 Target": None, "Q2 Achievement": None,
         "Q3 Target": None, "Q3 Achievement": None,
         "Q4 Target": None, "Q4 Achievement": None},
        {"Period": 2024, "Organization": "KNA", "Project Name": "REACH-KK",
         "Indicators": "Td Two Doses",
         "Q1 Target": None, "Q1 Achievement": None,
         "Q2 Target": None, "Q2 Achievement": None,
         "Q3 Target": None, "Q3 Achievement": None,
         "Q4 Target": None, "Q4 Achievement": None},
        {"Period": 2025, "Organization": "KNA", "Project Name": "REACH-KK",
         "Indicators": "Td Two Doses",
         "Q1 Target": 2725 / 6, "Q1 Achievement": None,
         "Q2 Target": 2725 / 6, "Q2 Achievement": None,
         "Q3 Target": 2725 / 6, "Q3 Achievement": None,
         "Q4 Target": 2725 / 6, "Q4 Achievement": None},
        {"Period": 2026, "Organization": "KNA", "Project Name": "REACH-KK",
         "Indicators": "Td Two Doses",
         "Q1 Target": 2900 / 3, "Q1 Achievement": None,
         "Q2 Target": 2900 / 3, "Q2 Achievement": None,
         "Q3 Target": None, "Q3 Achievement": None,
         "Q4 Target": None, "Q4 Achievement": None},
    ]
    return pd.DataFrame(rows)


# ── Write ──────────────────────────────────────────────────────────────────────
print(f"\nBuilding : 2025_indicators sheet...")
indicators_df = build_indicators_2025()
print(f"  Rows   : {len(indicators_df)}")

print(f"Building : Td_alod sheet...")
td_alod_df = build_td_alod()

print(f"Building : ALOD_cummu sheet...")
alod_cummu_df = build_alod_cummu()

print(f"Building : IDP sheet...")
idp_df = build_idp()

print(f"Building : Td2_indicator sheet...")
td2_df = build_td2_indicator()

print(f"\nWriting  : {DST}")
with pd.ExcelWriter(DST, engine="openpyxl") as writer:
  df.to_excel(writer, sheet_name="Child", index=False)
  td_df.to_excel(writer, sheet_name="Td", index=False)
  indicators_df.to_excel(writer, sheet_name="2025_indicators", index=False)
  td_alod_df.to_excel(writer, sheet_name="Td_alod", index=False)
  alod_cummu_df.to_excel(writer, sheet_name="ALOD_cummu", index=False)
  idp_df.to_excel(writer, sheet_name="IDP", index=False)
  td2_df.to_excel(writer, sheet_name="Td2_indicator", index=False)

  # Highlight duplicate children_code values in yellow on the Child sheet.
  child_ws = writer.book["Child"]
  yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
  cc_col_idx = df.columns.get_loc("children_code") + 1  # openpyxl is 1-based
  for row_idx in df.index[dup_mask]:
      child_ws.cell(row=int(row_idx) + 2, column=cc_col_idx).fill = yellow_fill

  # Highlight duplicate pw_code/PW_code values in yellow on the Td sheet.
  if td_code_col is not None:
      td_ws = writer.book["Td"]
      td_col_idx = td_df.columns.get_loc(td_code_col) + 1  # openpyxl is 1-based
      for row_idx in td_df.index[td_dup_mask]:
          td_ws.cell(row=int(row_idx) + 2, column=td_col_idx).fill = yellow_fill

print("Done.")
print(f"  Sheets written: Child, Td, 2025_indicators, Td_alod, ALOD_cummu, IDP, Td2_indicator")
print(f"  Child columns ({len(df.columns)}): {df.columns.tolist()}")
